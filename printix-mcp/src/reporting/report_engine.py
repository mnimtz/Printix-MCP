"""
Report Engine — HTML/CSV/JSON Ausgabe
======================================
Generiert Reports aus Query-Ergebnissen.

Unterstützte Formate (v1.0):
  html  — Vollständige HTML-Mail mit Branding
  csv   — Kommagetrennte Werte
  json  — Rohdaten als JSON

v1.1 (geplant):
  pdf   — WeasyPrint
  xlsx  — openpyxl
"""

import csv
import io
import json
import logging
from datetime import datetime
from typing import Any, Optional

logger = logging.getLogger(__name__)

try:
    from jinja2 import Environment, BaseLoader
    JINJA2_AVAILABLE = True
except ImportError:
    JINJA2_AVAILABLE = False
    logger.warning("jinja2 nicht installiert — HTML-Reports nicht verfügbar")


# ─── HTML Template ────────────────────────────────────────────────────────────

_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{{ title }}</title>
<style>
  body        { font-family: Arial, sans-serif; font-size: 13px; color: #333; margin: 0; padding: 0; background: #f5f5f5; }
  .wrapper    { max-width: 900px; margin: 20px auto; background: #fff; border-radius: 6px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,.12); }
  .header     { background: {{ primary_color }}; color: #fff; padding: 24px 32px; display: flex; align-items: flex-start; justify-content: space-between; gap: 20px; }
  .header-text { flex: 1; }
  .header h1  { margin: 0 0 4px; font-size: 22px; font-weight: 600; }
  .header p   { margin: 0; opacity: .85; font-size: 13px; }
  .header-logo { max-height: 56px; max-width: 160px; object-fit: contain; align-self: center; border-radius: 4px; }
  .content    { padding: 24px 32px; }
  .kpi-row    { display: flex; gap: 16px; margin-bottom: 24px; flex-wrap: wrap; }
  .kpi-card   { flex: 1; min-width: 140px; background: #f0f4ff; border-left: 4px solid {{ primary_color }}; padding: 14px 18px; border-radius: 4px; }
  .kpi-card .label { font-size: 11px; text-transform: uppercase; color: #666; letter-spacing: .5px; }
  .kpi-card .value { font-size: 24px; font-weight: 700; color: {{ primary_color }}; margin-top: 4px; }
  .kpi-card .sub   { font-size: 11px; color: #888; margin-top: 2px; }
  h2          { font-size: 15px; font-weight: 600; color: #333; margin: 24px 0 10px; border-bottom: 2px solid {{ primary_color }}; padding-bottom: 6px; }
  table       { width: 100%; border-collapse: collapse; font-size: 12px; margin-bottom: 16px; }
  th          { background: {{ primary_color }}; color: #fff; padding: 8px 12px; text-align: left; font-weight: 600; }
  td          { padding: 7px 12px; border-bottom: 1px solid #eee; }
  tr:nth-child(even) td { background: #f9f9f9; }
  tr:hover td { background: #eef4ff; }
  .delta-pos  { color: #d32f2f; font-weight: 600; }
  .delta-neg  { color: #2e7d32; font-weight: 600; }
  .delta-neu  { color: #888; }
  .badge      { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; }
  .badge-ok   { background: #e8f5e9; color: #2e7d32; }
  .badge-warn { background: #fff3e0; color: #e65100; }
  .footer     { background: #f5f5f5; padding: 14px 32px; font-size: 11px; color: #888; border-top: 1px solid #ddd; }
</style>
</head>
<body>
<div class="wrapper">
  <div class="header">
    <div class="header-text">
      <h1>{{ title }}</h1>
      <p>{{ company_name }} &nbsp;|&nbsp; Zeitraum: {{ period }} &nbsp;|&nbsp; Erstellt: {{ generated_at }}</p>
    </div>
    {%- if logo_url %}<img class="header-logo" src="{{ logo_url }}" alt="Logo">{% endif %}
  </div>
  <div class="content">

  {%- if kpis %}
  <div class="kpi-row">
    {%- for kpi in kpis %}
    <div class="kpi-card">
      <div class="label">{{ kpi.label }}</div>
      <div class="value">{{ kpi.value }}</div>
      {%- if kpi.sub %}<div class="sub">{{ kpi.sub }}</div>{% endif %}
    </div>
    {%- endfor %}
  </div>
  {%- endif %}

  {%- for section in sections %}
  <h2>{{ section.title }}</h2>
  {%- if section.rows %}
  <table>
    <thead><tr>
      {%- for col in section.columns %}<th>{{ col }}</th>{% endfor %}
    </tr></thead>
    <tbody>
      {%- for row in section.rows %}
      <tr>
        {%- for cell in row %}<td>{{ cell }}</td>{% endfor %}
      </tr>
      {%- endfor %}
    </tbody>
  </table>
  {%- else %}
  <p style="color:#888">Keine Daten für diesen Zeitraum.</p>
  {%- endif %}
  {%- endfor %}

  </div>
  <div class="footer">
    {{ footer_text }} &nbsp;|&nbsp; Printix MCP Report Engine &nbsp;|&nbsp; {{ generated_at }}
  </div>
</div>
</body>
</html>"""


# ─── Formatierungshilfen ──────────────────────────────────────────────────────

def _pdf_safe(text: str) -> str:
    """
    Ersetzt Unicode-Zeichen, die von fpdf2/Helvetica (Latin-1) nicht unterstützt werden.
    Verhindert FPDFUnicodeEncodingException bei Em-Dash, Sonderzeichen etc.
    """
    replacements = {
        "\u2013": "-",    # En-Dash –
        "\u2014": "-",    # Em-Dash —
        "\u2018": "'", "\u2019": "'",
        "\u201c": '"', "\u201d": '"',
        "\u2026": "...",
        "\u20ac": "EUR",
        "\u00b0": " Grad", "\u00b2": "2", "\u00b3": "3",
        "\u2122": "TM", "\u00ae": "(R)", "\u00a9": "(C)",
        "\u2022": "*", "\u25cf": "*",
    }
    for uni, asc in replacements.items():
        text = text.replace(uni, asc)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _fmt_num(v, decimals=0) -> str:
    if v is None:
        return "—"
    try:
        if decimals:
            return f"{float(v):,.{decimals}f}"
        return f"{int(v):,}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_pct(v) -> str:
    if v is None:
        return "—"
    return f"{float(v):.1f} %"


def _fmt_cost(v, currency="€") -> str:
    if v is None:
        return "—"
    return f"{float(v):,.2f} {currency}"


def _fmt_delta(v) -> str:
    if v is None:
        return "—"
    sign = "▲" if v > 0 else "▼"
    css  = "delta-pos" if v > 0 else "delta-neg"
    return f'<span class="{css}">{sign} {abs(v):.1f}%</span>'


# ─── Report-Generatoren ───────────────────────────────────────────────────────

def build_print_stats_report(
    rows: list[dict],
    period: str,
    layout: Optional[dict] = None,
    group_by: str = "day",
) -> dict[str, Any]:
    """Baut die Report-Datenstruktur für print_stats-Ergebnisse."""
    layout = layout or {}
    totals = {
        "total_jobs":   sum(r.get("total_jobs", 0) or 0 for r in rows),
        "total_pages":  sum(r.get("total_pages", 0) or 0 for r in rows),
        "color_pages":  sum(r.get("color_pages", 0) or 0 for r in rows),
        "bw_pages":     sum(r.get("bw_pages", 0) or 0 for r in rows),
        "duplex_pages": sum(r.get("duplex_pages", 0) or 0 for r in rows),
        "saved_sheets": sum(r.get("saved_sheets_duplex", 0) or 0 for r in rows),
    }
    color_pct  = totals["color_pages"] / max(totals["total_pages"], 1) * 100
    duplex_pct = totals["duplex_pages"] / max(totals["total_pages"], 1) * 100

    kpis = [
        {"label": "Seiten gesamt",    "value": _fmt_num(totals["total_pages"]),  "sub": f"{_fmt_num(totals['total_jobs'])} Aufträge"},
        {"label": "Farbseiten",       "value": _fmt_num(totals["color_pages"]),  "sub": _fmt_pct(color_pct)},
        {"label": "S/W-Seiten",       "value": _fmt_num(totals["bw_pages"]),     "sub": _fmt_pct(100 - color_pct)},
        {"label": "Duplex-Quote",     "value": _fmt_pct(duplex_pct),             "sub": f"{_fmt_num(totals['saved_sheets'])} Blätter gespart"},
    ]

    col_label = {"day": "Datum", "week": "Woche", "month": "Monat", "user": "Benutzer", "printer": "Drucker", "site": "Standort"}.get(group_by, "Periode")
    table_cols = [col_label, "Aufträge", "Seiten", "Farbe", "S/W", "Farb-%", "Duplex-%"]
    table_rows = []
    for r in rows:
        table_rows.append([
            str(r.get("period", "")).split("T")[0],
            _fmt_num(r.get("total_jobs")),
            _fmt_num(r.get("total_pages")),
            _fmt_num(r.get("color_pages")),
            _fmt_num(r.get("bw_pages")),
            _fmt_pct(r.get("color_pct")),
            _fmt_pct(r.get("duplex_pct")),
        ])

    return {
        "kpis": kpis,
        "sections": [{"title": "Druckvolumen Detail", "columns": table_cols, "rows": table_rows}],
    }


def build_cost_report_report(
    rows: list[dict],
    period: str,
    layout: Optional[dict] = None,
    currency: str = "€",
) -> dict[str, Any]:
    """Baut die Report-Datenstruktur für cost_report-Ergebnisse."""
    totals = {k: sum(r.get(k, 0) or 0 for r in rows)
              for k in ("total_pages", "color_pages", "bw_pages", "toner_cost_color",
                        "toner_cost_bw", "sheet_cost", "total_cost")}

    kpis = [
        {"label": "Gesamtkosten",  "value": _fmt_cost(totals["total_cost"], currency),   "sub": None},
        {"label": "Tonerkosten F", "value": _fmt_cost(totals["toner_cost_color"], currency), "sub": f"{_fmt_num(totals['color_pages'])} Farbseiten"},
        {"label": "Tonerkosten SW","value": _fmt_cost(totals["toner_cost_bw"], currency), "sub": f"{_fmt_num(totals['bw_pages'])} S/W-Seiten"},
        {"label": "Papierkosten",  "value": _fmt_cost(totals["sheet_cost"], currency),    "sub": f"{_fmt_num(totals['total_pages'])} Seiten"},
    ]

    table_cols = ["Periode", "Seiten", "Farbe", "S/W", "Toner F", "Toner SW", "Papier", "Gesamt"]
    table_rows = []
    for r in rows:
        table_rows.append([
            str(r.get("period", "")).split("T")[0],
            _fmt_num(r.get("total_pages")),
            _fmt_num(r.get("color_pages")),
            _fmt_num(r.get("bw_pages")),
            _fmt_cost(r.get("toner_cost_color"), currency),
            _fmt_cost(r.get("toner_cost_bw"), currency),
            _fmt_cost(r.get("sheet_cost"), currency),
            _fmt_cost(r.get("total_cost"), currency),
        ])

    return {
        "kpis": kpis,
        "sections": [{"title": "Kostenaufstellung nach Periode", "columns": table_cols, "rows": table_rows}],
    }


def build_top_users_report(rows: list[dict], period: str, currency: str = "€") -> dict[str, Any]:
    """Baut die Report-Datenstruktur für top_users-Ergebnisse."""
    table_cols = ["#", "Benutzer", "Abteilung", "Aufträge", "Seiten", "Farbe", "Farb-%", "Kosten"]
    table_rows = []
    for i, r in enumerate(rows, 1):
        table_rows.append([
            str(i),
            r.get("email", "—"),
            r.get("department") or "—",
            _fmt_num(r.get("total_jobs")),
            _fmt_num(r.get("total_pages")),
            _fmt_num(r.get("color_pages")),
            _fmt_pct(r.get("color_pct")),
            _fmt_cost(r.get("total_cost"), currency),
        ])
    kpis = []
    if rows:
        kpis = [
            {"label": "Aktivste Nutzer", "value": str(len(rows)),                             "sub": None},
            {"label": "Meiste Seiten",   "value": rows[0].get("email", "—"),                  "sub": _fmt_num(rows[0].get("total_pages"))},
        ]
    return {
        "kpis": kpis,
        "sections": [{"title": "Top Nutzer", "columns": table_cols, "rows": table_rows}],
    }


def build_top_printers_report(rows: list[dict], period: str, currency: str = "€") -> dict[str, Any]:
    """Baut die Report-Datenstruktur für top_printers-Ergebnisse."""
    table_cols = ["#", "Drucker", "Modell", "Standort", "Site", "Aufträge", "Seiten", "Farb-%", "Kosten"]
    table_rows = []
    for i, r in enumerate(rows, 1):
        table_rows.append([
            str(i),
            r.get("printer_name", "—"),
            r.get("model_name") or "—",
            r.get("location") or "—",
            r.get("site_name") or "—",
            _fmt_num(r.get("total_jobs")),
            _fmt_num(r.get("total_pages")),
            _fmt_pct(r.get("color_pct")),
            _fmt_cost(r.get("total_cost"), currency),
        ])
    return {
        "kpis": [],
        "sections": [{"title": "Top Drucker", "columns": table_cols, "rows": table_rows}],
    }


def build_trend_report(trend: dict, currency: str = "€") -> dict[str, Any]:
    """Baut die Report-Datenstruktur für trend-Ergebnisse."""
    p1, p2, delta = trend.get("period1", {}), trend.get("period2", {}), trend.get("delta", {})

    table_cols = ["Kennzahl", "Periode 1", "Periode 2", "Veränderung"]
    metrics = [
        ("Seiten gesamt",    "total_pages",    _fmt_num),
        ("Farbseiten",       "color_pages",    _fmt_num),
        ("S/W-Seiten",       "bw_pages",       _fmt_num),
        ("Aufträge",         "total_jobs",     _fmt_num),
        ("Aktive Nutzer",    "active_users",   _fmt_num),
        ("Gesamtkosten",     "total_cost",     lambda v: _fmt_cost(v, currency)),
    ]
    table_rows = []
    for label, key, fmt in metrics:
        table_rows.append([
            label,
            fmt(p1.get(key)),
            fmt(p2.get(key)),
            _fmt_delta(delta.get(key)),
        ])

    p1_range = f"{p1.get('start','?')} – {p1.get('end','?')}"
    p2_range = f"{p2.get('start','?')} – {p2.get('end','?')}"

    return {
        "kpis": [
            {"label": "Seiten (Periode 1)", "value": _fmt_num(p1.get("total_pages")), "sub": p1_range},
            {"label": "Seiten (Periode 2)", "value": _fmt_num(p2.get("total_pages")), "sub": p2_range},
            {"label": "Veränderung Seiten", "value": f"{delta.get('total_pages', '—')} %", "sub": None},
            {"label": "Veränderung Kosten", "value": f"{delta.get('total_cost', '—')} %",  "sub": None},
        ],
        "sections": [{"title": "Periodischer Vergleich", "columns": table_cols, "rows": table_rows}],
    }


# ─── Ausgabe-Renderer ─────────────────────────────────────────────────────────

def render_html(
    title: str,
    period: str,
    report_data: dict[str, Any],
    layout: Optional[dict] = None,
) -> str:
    """Rendert den Report als vollständige HTML-Seite."""
    if not JINJA2_AVAILABLE:
        # Fallback: einfaches HTML ohne Jinja2
        return _plain_html_fallback(title, period, report_data, layout)

    layout = layout or {}
    env = Environment(loader=BaseLoader())
    template = env.from_string(_HTML_TEMPLATE)
    return template.render(
        title=title,
        period=period,
        company_name=layout.get("company_name", ""),
        primary_color=layout.get("primary_color", "#0078D4"),
        footer_text=layout.get("footer_text", ""),
        logo_url=layout.get("logo_url", ""),
        generated_at=datetime.now().strftime("%d.%m.%Y %H:%M"),
        kpis=report_data.get("kpis", []),
        sections=report_data.get("sections", []),
    )


def _plain_html_fallback(title, period, report_data, layout) -> str:
    lines = [f"<html><body><h1>{title}</h1><p>{period}</p>"]
    for section in report_data.get("sections", []):
        lines.append(f"<h2>{section['title']}</h2><table border='1'>")
        lines.append("<tr>" + "".join(f"<th>{c}</th>" for c in section["columns"]) + "</tr>")
        for row in section["rows"]:
            lines.append("<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>")
        lines.append("</table>")
    lines.append("</body></html>")
    return "\n".join(lines)


def render_csv(rows: list[dict]) -> str:
    """Rendert rohe Query-Ergebnisse als CSV."""
    if not rows:
        return ""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()), extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({k: str(v) if v is not None else "" for k, v in row.items()})
    return output.getvalue()


def render_json(data: Any) -> str:
    """Serialisiert Daten als formatiertes JSON."""
    return json.dumps(data, indent=2, ensure_ascii=False, default=str)


def render_pdf(title: str, period: str, report_data: dict, layout: Optional[dict] = None) -> bytes:
    """
    Erzeugt einen PDF-Report aus den Report-Daten.
    Verwendet fpdf2 mit Helvetica (Latin-1) — alle Nicht-Latin1-Zeichen werden durch _pdf_safe() ersetzt.
    """
    try:
        from fpdf import FPDF
    except ImportError:
        logger.error("fpdf2 nicht installiert — PDF-Generierung nicht verfügbar")
        return b""

    layout = layout or {}
    primary_color = layout.get("primary_color", "#1a73e8")

    # Farbe parsen (Hex → RGB)
    def hex_to_rgb(h: str):
        h = h.lstrip("#")
        if len(h) == 3:
            h = "".join(c * 2 for c in h)
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

    try:
        r, g, b = hex_to_rgb(primary_color)
    except Exception:
        r, g, b = 26, 115, 232

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Header
    pdf.set_fill_color(r, g, b)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 12, _pdf_safe(title), ln=True, fill=True, align="L")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 7, _pdf_safe(f"Zeitraum: {period}  |  Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"), ln=True, fill=True, align="L")
    pdf.ln(4)

    # KPIs
    kpis = report_data.get("kpis", [])
    if kpis:
        pdf.set_fill_color(240, 244, 255)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 6, "Kennzahlen", ln=True)
        pdf.ln(2)
        for kpi in kpis:
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(r, g, b)
            pdf.cell(0, 7, _pdf_safe(f"  {kpi.get('label','')}: {kpi.get('value','')}"), ln=True, fill=True)
            if kpi.get("sub"):
                pdf.set_font("Helvetica", "", 9)
                pdf.set_text_color(120, 120, 120)
                pdf.cell(0, 5, _pdf_safe(f"    {kpi['sub']}"), ln=True)
        pdf.ln(4)

    # Sections / Tabellen
    pdf.set_text_color(0, 0, 0)
    for section in report_data.get("sections", []):
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(r, g, b)
        pdf.cell(0, 8, _pdf_safe(section.get("title", "")), ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(1)

        columns = section.get("columns", [])
        rows = section.get("rows", [])
        if not rows:
            pdf.set_font("Helvetica", "I", 9)
            pdf.set_text_color(150, 150, 150)
            pdf.cell(0, 6, "Keine Daten.", ln=True)
            pdf.ln(3)
            continue

        # Spaltenbreiten gleichmäßig verteilen
        page_w = pdf.w - 2 * pdf.l_margin
        col_w = page_w / max(len(columns), 1)

        # Tabellenkopf
        pdf.set_fill_color(r, g, b)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        for col in columns:
            pdf.cell(col_w, 7, _pdf_safe(str(col)), border=0, fill=True)
        pdf.ln()

        # Tabellenzeilen
        pdf.set_font("Helvetica", "", 8)
        for i, row in enumerate(rows):
            if i % 2 == 0:
                pdf.set_fill_color(249, 249, 249)
            else:
                pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(0, 0, 0)
            for cell in row:
                pdf.cell(col_w, 6, _pdf_safe(str(cell)), border=0, fill=True)
            pdf.ln()
        pdf.ln(4)

    return bytes(pdf.output())


def render_xlsx(title: str, period: str, report_data: dict, layout: Optional[dict] = None) -> bytes:
    """
    Erzeugt einen Excel-Report (XLSX) aus den Report-Daten.
    Verwendet openpyxl.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.error("openpyxl nicht installiert — XLSX-Generierung nicht verfügbar")
        return b""

    layout = layout or {}
    primary_hex = layout.get("primary_color", "#1a73e8").lstrip("#")
    if len(primary_hex) == 3:
        primary_hex = "".join(c * 2 for c in primary_hex)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    header_fill = PatternFill("solid", fgColor=primary_hex)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    kpi_font    = Font(bold=True, size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    row_idx = 1

    # Titel
    ws.cell(row_idx, 1, title)
    ws.cell(row_idx, 1).font = Font(bold=True, size=14)
    row_idx += 1
    ws.cell(row_idx, 1, f"Zeitraum: {period}  |  Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    ws.cell(row_idx, 1).font = Font(italic=True, size=9, color="888888")
    row_idx += 2

    # KPIs
    kpis = report_data.get("kpis", [])
    if kpis:
        ws.cell(row_idx, 1, "Kennzahlen")
        ws.cell(row_idx, 1).font = Font(bold=True, size=11)
        row_idx += 1
        for kpi in kpis:
            ws.cell(row_idx, 1, kpi.get("label", ""))
            ws.cell(row_idx, 2, kpi.get("value", ""))
            ws.cell(row_idx, 2).font = kpi_font
            if kpi.get("sub"):
                ws.cell(row_idx, 3, kpi["sub"])
                ws.cell(row_idx, 3).font = Font(size=9, color="888888")
            row_idx += 1
        row_idx += 1

    # Sections / Tabellen
    for section in report_data.get("sections", []):
        ws.cell(row_idx, 1, section.get("title", ""))
        ws.cell(row_idx, 1).font = Font(bold=True, size=12)
        row_idx += 1

        columns = section.get("columns", [])
        rows    = section.get("rows", [])

        if columns:
            for col_idx, col in enumerate(columns, start=1):
                cell = ws.cell(row_idx, col_idx, col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left")
            row_idx += 1

        alt_fill = PatternFill("solid", fgColor="F9F9F9")
        for i, row in enumerate(rows):
            for col_idx, cell_val in enumerate(row, start=1):
                cell = ws.cell(row_idx, col_idx, cell_val)
                cell.border = thin_border
                if i % 2 == 0:
                    cell.fill = alt_fill
            row_idx += 1
        row_idx += 2

    # Spaltenbreiten anpassen
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_report(
    query_type: str,
    data: Any,
    period: str,
    layout: Optional[dict] = None,
    output_formats: Optional[list[str]] = None,
    currency: str = "€",
) -> dict[str, str]:
    """
    Erzeugt alle angeforderten Ausgabeformate für einen Report.

    Returns:
        Dict {format: content} — z.B. {"html": "<!DOCTYPE html>...", "csv": "col1,col2\n..."}
    """
    output_formats = output_formats or ["html"]
    layout = layout or {}

    # Report-Datenstruktur bauen
    if query_type == "print_stats":
        report_data = build_print_stats_report(data, period, layout)
        title = "Druckvolumen-Report"
    elif query_type == "cost_report":
        report_data = build_cost_report_report(data, period, layout, currency)
        title = "Kostenaufstellung"
    elif query_type == "top_users":
        report_data = build_top_users_report(data, period, currency)
        title = "Top Nutzer Report"
    elif query_type == "top_printers":
        report_data = build_top_printers_report(data, period, currency)
        title = "Top Drucker Report"
    elif query_type == "trend":
        report_data = build_trend_report(data, currency)
        title = "Trend-Vergleich"
    else:
        report_data = {"kpis": [], "sections": [{"title": "Ergebnis", "columns": list(data[0].keys()) if data else [], "rows": [[str(v) for v in r.values()] for r in data] if data else []}]}
        title = query_type.replace("_", " ").title()

    # Gewünschte Formate generieren
    title_full = layout.get("company_name", "Printix") + " – " + title if layout.get("company_name") else title
    results = {}
    if "html" in output_formats:
        results["html"] = render_html(title_full, period, report_data, layout)
    if "csv" in output_formats:
        raw = data if isinstance(data, list) else []
        _csv = render_csv(raw)
        results["csv"] = _csv if _csv else f'Zeitraum,Hinweis\r\n"{period}",Keine Daten im abgefragten Zeitraum\r\n'
    if "json" in output_formats:
        results["json"] = render_json(data)
    if "pdf" in output_formats:
        try:
            results["pdf"] = render_pdf(title_full, period, report_data, layout)
        except Exception as e:
            logger.error("PDF-Generierung fehlgeschlagen: %s", e, exc_info=True)
            results["pdf_error"] = str(e)
    if "xlsx" in output_formats:
        try:
            results["xlsx"] = render_xlsx(title_full, period, report_data, layout)
        except Exception as e:
            logger.error("XLSX-Generierung fehlgeschlagen: %s", e, exc_info=True)
            results["xlsx_error"] = str(e)

    return results
